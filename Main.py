from tkinter import *
from openpyxl import Workbook
from openpyxl import load_workbook
import datetime


#from mcp3208 import MCP3208

# Set some Globabl Limits
# Set your warning conditions for high and low voltage here
MinMainVolt = 27
WarnMainVolt = 28
MaxMainVolt = 31

MinAuxVolt = 11
WarnAuxVolt = 12
MaxAuxVolt = 15

record_interval=1000

#scale factor to correct counts to volts
Scaleone = .013854


class App():
    def calcamphour(self,current,amphr):
        #import the the global variable needed
        global record_interval

        #calculate the amp hours.  Record_interval is in milliseconds.  Convert it to seconds by divife by 1000 and then to hours by multiplying by 1/3600
        amphr=amphr-(current*(record_interval/1000)*(1/3600))
        #sets the amphour variable and updates screen
        self.amphour.set(round(amphr,3))


    #records data to an excel spread sheet
    def excelrecord(self,drivedate, dataset):

        try:
            wb = load_workbook('History.xlsx')  # attemps to open the history excel file
        except:
            wb = Workbook()  # creates and empty excel workbook if histortory is not found

        #creates the name for the worksheet
        WorkSheetName=drivedate


        try:
            ws = wb['%s' %(WorkSheetName)]  # attemps to open the existing worksheete

        except:
            ws = wb.create_sheet(WorkSheetName) # creates and empty excel worksheet if the worksheet does not exist
            ws['A1']="Time"
            ws['B1']="Driver"
            ws['C1']="Main Volt"
            ws['D1']="Aux Volt"
            ws['E1']="Current"
            ws['F1']="Amp Hours"

        #add the new data to the worksheet
        ws.append(dataset)

        #save the worksheet
        wb.save('History.xlsx')


    #function that reocrds data from sources
    def recorddata(self):

        # import the gobal
        global MinMainVolt
        global WarnMainVolt
        global MaxMainVolt
        global MinAuxVolt
        global WarnAuxVolt
        global MaxAuxVolt
        global Scaleone
        global record_interval

        #gets the default background color so we can use it later
        defaultbg = root.cget('bg')

        #create the analog to digital converter
#        adc = MCP3208()

        #read the analog to digital converter
        mainreading = 1400 #adc.read(7)
        auxreading = 12.5 #adc.read(0)
        currentreaading = 50


        # Set the back ground color depending on the voltage
        self.mainvolt.set(mainreading * Scaleone) # convert the adc value from counts to volts, set the variable, and display on screen
        if self.mainvolt.get() < MinMainVolt or self.mainvolt.get() > MaxMainVolt:
            self.mainvoltlabel.config(bg='red')
        elif self.mainvolt.get() < WarnMainVolt:
            self.mainvoltlabel.config(bg='yellow')
        else:
            self.mainvoltlabel.config(bg=defaultbg)

        self.auxvolt.set(auxreading) # convert the adc value from counts to volts, set the variable, and display on screen
        if self.auxvolt.get() < MinMainVolt or self.auxvolt.get() > MaxAuxVolt:
            self.auxvoltlabel.config(bg='red')
        elif self.auxvolt.get() < WarnMainVolt:
            self.auxvoltlabel.config(bg='yellow')
        else:
            self.auxvoltlabel.config(bg=defaultbg)

        #calculate AmpHours
        self.calcamphour(currentreaading,self.amphour.get())


        #create a dataset for excel
        dataset=[datetime.datetime.now().strftime("%I:%M:%S %p"), self.drivername.get(),mainreading,auxreading, currentreaading,self.amphour.get()]
        self.excelrecord(self.drivedate.get(),dataset)

        # causes the program to recall the record segment every 1000 ms
        root.after(record_interval, lambda: self.recorddata())

    def setamphour(self):

        #define the window for amp hour reset
        top = Toplevel(width=90)
        top.title("Set Amp Hours")

        #message
        msg = Message(top, text="Set Amp Hours")
        msg.pack()
        ah=Entry(top)
        ah.pack()

        #Buttons for amphour set window
        button1 = Button(top, text="Set",command=lambda: self.amphour.set((ah.get())))
        button1.pack()

        button = Button(top, text="Dismiss", command=top.destroy)
        button.pack()


    #set up the Graphic user interface
    def __init__(self, master):

        frame = Frame(master)
        frame.pack()
        master.title("Solar Car Telemetry")
        textsize=40


        # declare some variables
        self.drivername = StringVar()
        self.drivedate = StringVar()
        self.mainvolt = DoubleVar()
        self.auxvolt = DoubleVar()
        self.current = DoubleVar()
        self.amphour = DoubleVar()

        # set some defaults to initialize variables
        self.main = 0
        self.aux = 0
        self.cur = 50.0
        self.amphr = 100

        # set some variables
        self.mainvolt.set(self.main)
        self.auxvolt.set(self.aux)
        self.current.set(self.cur)
        self.amphour.set(self.amphr)

        # define the labels
        self.drivernamelabel = Label(frame, text="Driver:",font=("Helvetica", textsize))
        self.drivedatelabel = Label(frame, text="Date:",font=("Helvetica", textsize))
        self.mainvoltlabel = Label(frame, text="Main Voltage:",font=("Helvetica", textsize))
        self.auxvoltlabel = Label(frame, text="Aux Voltage:",font=("Helvetica", textsize))
        self.currentlabel = Label(frame, text="Current:",font=("Helvetica", textsize))
        self.amphourlabel = Label(frame, text="AmpHour:",font=("Helvetica", textsize))

        self.drivername = Entry(frame,font=("Helvetica", textsize))
        self.drivedate = Entry(frame,font=("Helvetica", textsize))
        self.mainvoltvalue = Label(frame, textvariable=self.mainvolt,font=("Helvetica", textsize))
        self.auxvoltlvalue = Label(frame, textvariable=self.auxvolt,font=("Helvetica", textsize))
        self.currentvalue = Label(frame, textvariable=self.current,font=("Helvetica", textsize))
        self.amphourvalue = Label(frame, textvariable=self.amphour,font=("Helvetica", textsize))

        # place labels in grid
        self.drivernamelabel.grid(row=0, column=0)
        self.drivedatelabel.grid(row=1, column=0)
        self.mainvoltlabel.grid(row=2, column=0)
        self.auxvoltlabel.grid(row=3, column=0)
        self.currentlabel.grid(row=4, column=0)
        self.amphourlabel.grid(row=5, column=0)

        self.drivername.grid(row=0, column=1)
        self.drivedate.grid(row=1, column=1)
        self.mainvoltvalue.grid(row=2, column=1)
        self.auxvoltlvalue.grid(row=3, column=1)
        self.currentvalue.grid(row=4, column=1)
        self.amphourvalue.grid(row=5, column=1)

        # define the buttons
        self.button = Button(frame, text="Quit", width=15, command=frame.quit)
        self.record = Button(frame, text="Record", width=15, command=lambda: self.recorddata())

        self.record.grid(row=6, column=0, sticky='n')
        self.button.grid(row=6, column=1, sticky='n')

        self.setamphr = Button(frame, text="Set Amp Hours", width=15, command=lambda: self.setamphour())
        self.setamphr.grid(row=7, column=1, sticky='n')

#established that this program is TKinter GUI
root = Tk()
#creates the specific app
app = App(root)
#starts the main loop
root.mainloop()